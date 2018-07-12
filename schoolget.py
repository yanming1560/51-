import time,random,os
from selenium import webdriver
from openpyxl import load_workbook as lw

def xls_list(str,num):      #基本信息写入excel
    all_text=str.split('\n')
    name=all_text[0]
    zone,pro1,pro2,add,tel='none','none','none','none','none'
    for i in all_text:
        if ('地区' in i):
            zone=i.split(':')[1]
        elif ('属性' in i):
            pro1=i.split(':')[1]
        elif ('性质' in i):
            pro2 = i.split(':')[1]
        elif ('地址' in i):
            add = i.split(':')[1]
        elif ('电话' in i):
            tel = i.split(':')[1]
    ws.cell(row=num + 1, column=1).value = num
    ws.cell(row=num+1, column=2).value = name
    ws.cell(row=num+1, column=3).value = zone
    ws.cell(row=num+1, column=4).value = pro1
    ws.cell(row=num+1, column=5).value = pro2
    ws.cell(row=num+1, column=6).value = add
    ws.cell(row=num+1, column=7).value = tel

def introduce(driver,i,num):   #输出收费情况
    name=i.text.split('\n')[0]
    time.sleep(random.uniform(1, 2))
    driver.find_element_by_link_text(name).click()      #点击进入学校页面
    winall = driver.window_handles          #得到所有窗口
    driver.switch_to.window(winall[-1])      #选择新打开的窗口为当前窗口

    try:            #尝试找到费用
        b = driver.find_element_by_class_name('school_y_sf')
        if ('[点击查看完整收费]' in b.text):        #如果有这个按钮，费用需要展开
            time.sleep(random.uniform(1, 2))
            driver.find_element_by_link_text('[点击查看完整收费]').click()      #打开完整费用页面
            winall = driver.window_handles  # 得到所有窗口
            driver.switch_to.window(winall[-1])  # 选择新打开的窗口为当前窗口
            c=driver.find_element_by_class_name('nr_m')     #选择费用元素
            with open('youeryuan/' + str(num) + '/charge.txt', 'w') as f:  # 学校收费情况
                f.write(c.text)
            time.sleep(random.uniform(1, 2))
            driver.close()      #关闭完整费用页面窗口
            winall = driver.window_handles  # 得到所有窗口
            driver.switch_to.window(winall[-1])  # 将当前窗口改为学校页面
        else:
            with open('youeryuan/' + str(num) + '/charge.txt', 'w') as f:  # 学校收费情况
                f.write(b.text)
    except:
        print(num,'none charge information')

    try:        #尝试找到招生简章
        driver.find_element_by_link_text('[点击查看完整招生简章]').click()    #进去招生简章详情页
        winall = driver.window_handles  # 得到所有窗口
        driver.switch_to.window(winall[-1])  # 选择新打开的窗口为当前窗口
        c = driver.find_element_by_class_name('nr_m')  # 选择招生简章元素
        with open('youeryuan/' + str(num) + '/introduce.txt', 'w') as f:  # 学校招生简章
            f.write(c.text)
        time.sleep(random.uniform(1, 2))
        driver.close()  # 关闭招生简章页面窗口
        winall = driver.window_handles  # 得到所有窗口
        driver.switch_to.window(winall[-1])  # 将当前窗口改为学校页面
    except:
        print(num,'none introduce')

    time.sleep(random.uniform(1, 2))
    driver.close()      #关闭学校页面窗口
    winall = driver.window_handles  # 得到所有窗口
    if len(winall)>1:
        driver.switch_to.window(winall[-1])
        driver.close()
        driver.switch_to.window(winall[0])
    elif len(winall)==1:
        driver.switch_to.window(winall[0])


def get_school(driver,num):     #单个页面的采集
    a = driver.find_elements_by_class_name('reply_box')     #reply_box标注页面内的学校
    for i in a:
        os.makedirs('youeryuan/' + str(num))        #建立学校编号文件夹
        with open('youeryuan/' + str(num) + '/base.txt', 'w') as f:    #学校基本信息文件
            f.write(i.text)
        #xls_list(i.text,num)        #学校基本信息写入表格
        introduce(driver,i,num)       #学校简介和收费
        print('school ',num, ' finished!')
        num+=1
    time.sleep(random.uniform(1, 2))
    driver.find_element_by_link_text('下一页').click()     #进入下一个页面
    return num      #返回下一个的编号值

def title(ws):      #excel表格的表头
    ws.cell(row=1, column=1).value = '编号'
    ws.cell(row=1, column=2).value = '名称'
    ws.cell(row=1, column=3).value = '地区'
    ws.cell(row=1, column=4).value = '属性'
    ws.cell(row=1, column=5).value = '性质'
    ws.cell(row=1, column=6).value = '地址'
    ws.cell(row=1, column=7).value = '电话'

if __name__=='__main__':
    fm=lw('menu_yey.xlsx')      #打开幼儿园目录excel表格
    ws=fm['Sheet1']             #选中sheet1进行操作
    title(ws)                   #设置表头

    site = 'http://xuexiao.51sxue.com/slist/?t=1&areaCodeS=31&page=1'       #目标url
    driver = webdriver.Chrome()         #chrome浏览器driver
    driver.get(site)                    #打开url
    time.sleep(random.uniform(1, 2))        #随机等待1-2s
    driver.maximize_window()            #窗口最大化
    num=1               #初始学校编号为1
    for i in range(1,50):        #爬取页面从x到x-1
        try:
            num=get_school(driver,num)      #爬取主程序
            print('page ',i,' finished!')
        except:
            print('someting went wrong at page ',i)
            break
    #fm.save('menu_yey.xlsx')        #目录excel表格
    print('next number is ',num)
