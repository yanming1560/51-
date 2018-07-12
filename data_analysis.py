from openpyxl import load_workbook as lw
import matplotlib.pyplot as plt
import numpy as np

class school():
    def __init__(self,name,zone,lev,pro,add):
        self.name=name
        self.zone=zone
        self.lev=lev
        self.pro=pro
        self.add=add



def check_zone(t_zon):
    f = lw('幼儿园名录.xlsx')  # 打开目录文档
    wb = f['Sheet1']  # 找到工作表
    for i in range(1,1814):     #定义所有的学校变量，如s1,s23
        vars()['s'+str(i)]=school(wb.cell(row=i+1, column=2).value,wb.cell(row=i+1, column=3).value,wb.cell(row=i+1, column=4).value,wb.cell(row=i+1, column=5).value,wb.cell(row=i+1, column=6).value)
    f.close()
    total,gong,gshi,gyi,ger,gsan,gpt,sshi,syi,ser,ssan,spt,si=0,0,0,0,0,0,0,0,0,0,0,0,0
    for i in range(1,1814):     #判断地区t_zon的数据
        if (t_zon in str(vars()['s'+str(i)].zone)) or (t_zon in str(vars()['s'+str(i)].add)):
            total+=1
            if ('公办' in str(vars()['s'+str(i)].pro)):
                gong+=1
                if ('示范' in str(vars()['s'+str(i)].lev)):
                    gshi+=1
                if ('一级' in str(vars()['s'+str(i)].lev)):
                    gyi+=1
                if ('二级' in str(vars()['s'+str(i)].lev)):
                    ger+=1
                if ('三级' in str(vars()['s'+str(i)].lev)):
                    gsan+=1
                if ('普通' in str(vars()['s'+str(i)].lev)):
                    gpt+=1
            else:
                si+=1
                if ('示范' in str(vars()['s'+str(i)].lev)):
                    sshi+=1
                if ('一级' in str(vars()['s'+str(i)].lev)):
                    syi+=1
                if ('二级' in str(vars()['s'+str(i)].lev)):
                    ser+=1
                if ('三级' in str(vars()['s'+str(i)].lev)):
                    ssan+=1
                if ('普通' in str(vars()['s'+str(i)].lev)):
                    spt+=1
    print(t_zon,'ok')
    return [gong,gshi,gyi,ger,gsan,gpt,si,sshi,syi,ser,ssan,spt]

if __name__=='__main__':        #对比所有区域幼儿园等级一级数量，并且画图
#def hhaa():
    all_zone=['普陀','徐汇','浦东','长宁','黄浦','青浦','杨浦','崇明','静安','松江','闵行','奉贤','金山']
    all_zone_en=['PuTuo','XuHui','PuDong','ChangNing','HuangPu','QingPu','YangPu','ChongMing',
                 'JingAn','SongJiang','MinHang','FengXian','JinShan']
    all_data=[]
    for i in range(5):
        vars()['g'+str(i)]=[]
        vars()['s'+str(i)]=[]
    for i in all_zone:
        all_data.append(check_zone(i))      #得到每一个城市的数据
    for i in range(13):
        for j in range(5):
            vars()['g' + str(j)].append(all_data[i][j+1])
            vars()['s' + str(j)].append(all_data[i][j+7])
    #print(g0,g1,g2,g3,g4,s0,s1,s2,s3,s4)

    ind=np.arange(13)*3
    ind2=[ind[i]+1 for i in range(13)]
    width=1

    p1=plt.bar(ind,g4,width)
    gb=g4
    p2=plt.bar(ind,g3,width,bottom=gb)
    gb = [g4[i]+g3[i] for i in range(13)]
    p3 = plt.bar(ind, g2, width, bottom=gb)
    gb = [g4[i]+g3[i]+g2[i] for i in range(13)]
    p4 = plt.bar(ind, g1, width, bottom=gb)
    gb = [g4[i]+g3[i]+g2[i]+g1[i] for i in range(13)]
    p5 = plt.bar(ind, g0, width, bottom=gb)

    p6 = plt.bar(ind2, s4, width)
    sb = s4
    p7 = plt.bar(ind2, s3, width, bottom=sb)
    sb = [s4[i] + s3[i] for i in range(13)]
    p8 = plt.bar(ind2, s2, width, bottom=sb)
    sb = [s4[i] + s3[i] + s2[i] for i in range(13)]
    p9 = plt.bar(ind2, s1, width, bottom=sb)
    sb = [s4[i] + s3[i] + s2[i] + s1[i] for i in range(13)]
    p10 = plt.bar(ind2, s0, width, bottom=sb)

    plt.ylabel('Number')
    plt.title('school distribution')
    plt.xticks(ind,all_zone_en )
    plt.legend((p1[0], p2[0],p3[0],p4[0],p5[0],p6[0], p7[0],p8[0],p9[0],p10[0]),
               ('GPuTong', 'GSanJi', 'GErJi', 'GYiJi', 'GShiFan','SPuTong', 'SSanJi', 'SErJi', 'SYiJi', 'SShiFan'))

    plt.show()

#if __name__=='__main__':       #单独输出一个区域各等级幼儿园数量
    #print(check_zone('浦东'))
